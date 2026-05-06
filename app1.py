import os
import json
import math
import subprocess
import tempfile
import shutil
from collections import defaultdict, Counter
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__, static_folder="static")
CORS(app)

TANK_DB_FILE = "tank_db.json"
TANK_DB = {}
if os.path.exists(TANK_DB_FILE):
    with open(TANK_DB_FILE, encoding="utf-8") as f:
        raw_db = json.load(f)
        TANK_DB = {int(k): v for k, v in raw_db.items()}

TYPE_LABELS = {
    "heavyTank":  "HT",
    "mediumTank": "MT",
    "lightTank":  "LT",
    "AT-SPG":     "TD",
}


def run_parser(file_path):
    result = subprocess.run(
        ["wotbreplay-inspector", "battle-results", file_path],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        return None
    try:
        return json.loads(result.stdout)
    except Exception:
        return None


def make_player_store():
    return {
        "nickname": "unknown",
        "battles": 0,
        "damage": 0,
        "frags": 0,
        "shots": 0,
        "hits": 0,
        "pens": 0,
        "assist": 0,
        "blocked": 0,
        "enemies_damaged": 0,
        "iPoints": 0,
        "sPoints": 0,
        "tank_battles": Counter(),
        "type_battles": Counter(),
    }


def calc_stats(acc_id, p):
    b = p["battles"]
    if b == 0:
        return None

    ADR     = p["damage"] / b
    KPR     = p["frags"] / b
    DE      = p["enemies_damaged"] / b
    Assist  = p["assist"] / b
    Blocked = p["blocked"] / b
    AccH    = p["hits"] / p["shots"] if p["shots"] else 0
    AccP    = p["pens"] / p["hits"]  if p["hits"]  else 0
    iPoints = p["iPoints"] / b
    sPoints = p["sPoints"] / b

    Firepower = ((100 + ADR) * (1 + KPR) ** (1/7) - 777) / 20
    AIM       = ((1 + AccH) * (1 + AccP) - 0.9) / 0.029
    Support   = ((1 + DE)**2 * (200 + Assist)**2 * (400 + Blocked)) ** (1/3) / 19
    Supremacy = math.sqrt(max(0, 40 + iPoints + sPoints)) / 0.13
    BPR       = (1/76) * ((17*Firepower + 3*AIM + 2*Support + 3*Supremacy) / 25)

    main_tank = "unknown"
    if p["tank_battles"]:
        top_id = p["tank_battles"].most_common(1)[0][0]
        main_tank = TANK_DB.get(top_id, {}).get("name", "unknown")

    return {
        "account_id": acc_id,
        "nickname":   p["nickname"],
        "battles":    b,
        "HT":         p["type_battles"].get("HT", 0),
        "MT":         p["type_battles"].get("MT", 0),
        "LT":         p["type_battles"].get("LT", 0),
        "TD":         p["type_battles"].get("TD", 0),
        "main_tank":  main_tank,
        "ADR":        round(ADR, 2),
        "frags":      p["frags"],
        "KPR":        round(KPR, 2),
        "DE":         round(DE, 2),
        "assist":     round(Assist, 2),
        "blocked":    round(Blocked, 2),
        "shots":      p["shots"],
        "hits":       p["hits"],
        "pens":       p["pens"],
        "AccH":       round(AccH * 100, 2),
        "AccP":       round(AccP * 100, 2),
        "iPoints":    round(iPoints, 2),
        "sPoints":    round(sPoints, 2),
        "Firepower":  round(Firepower, 2),
        "AIM":        round(AIM, 2),
        "Support":    round(Support, 2),
        "Supremacy":  round(Supremacy, 2),
        "BPR":        round(BPR, 2),
    }


def bpr_adr_sort(rows):
    """Sort by BPR desc, then ADR desc as tiebreaker."""
    return sorted(rows, key=lambda x: (-x["BPR"], -x["ADR"]))


def process_replays(folder_path):
    # separate stores per team label: "our" and "enemy"
    our_players   = defaultdict(make_player_store)
    enemy_players = defaultdict(make_player_store)

    errors    = []
    processed = 0

    for file in sorted(os.listdir(folder_path)):
        if not file.endswith(".wotbreplay"):
            continue

        full_path = os.path.join(folder_path, file)
        data = run_parser(full_path)
        if not data:
            errors.append(file)
            continue

        processed += 1

        author_id = data.get("author", {}).get("account_id")
        our_team  = None

        player_info_map = {}
        for p in data.get("players", []):
            acc_id = p.get("account_id")
            info   = p.get("info", {})
            if acc_id:
                player_info_map[acc_id] = {
                    "nickname": info.get("nickname", "unknown"),
                    "team":     info.get("team"),
                }
                if acc_id == author_id:
                    our_team = info.get("team")

        if our_team is None:
            errors.append(f"{file} (team not found)")
            continue

        our_ids   = {a for a, i in player_info_map.items() if i["team"] == our_team}
        enemy_ids = {a for a, i in player_info_map.items() if i["team"] != our_team}

        for p in data.get("player_results", []):
            info   = p.get("info", {})
            acc_id = info.get("account_id")
            if not acc_id:
                continue

            # choose store
            if acc_id in our_ids:
                store = our_players
            elif acc_id in enemy_ids:
                store = enemy_players
            else:
                continue

            pl = store[acc_id]
            if acc_id in player_info_map:
                pl["nickname"] = player_info_map[acc_id]["nickname"]

            pl["battles"]         += 1
            pl["damage"]          += info.get("damage_dealt", 0)
            pl["frags"]           += info.get("n_enemies_destroyed", 0)
            pl["shots"]           += info.get("n_shots", 0)
            pl["hits"]            += info.get("n_hits_dealt", 0)
            pl["pens"]            += info.get("n_penetrations_dealt", 0)
            pl["assist"]          += info.get("damage_assisted_1", 0) + info.get("damage_assisted_2", 0)
            pl["blocked"]         += info.get("damage_blocked", 0)
            pl["enemies_damaged"] += info.get("n_enemies_damaged", 0)

            earned   = info.get("victory_points_earned", 0)
            captured = info.get("victory_points_seized", 0)
            pl["iPoints"] += (earned - captured)
            pl["sPoints"] += captured

            tank_id = info.get("tank_id")
            if tank_id:
                pl["tank_battles"][tank_id] += 1
                wg_type = TANK_DB.get(tank_id, {}).get("type", "unknown")
                label   = TYPE_LABELS.get(wg_type)
                if label:
                    pl["type_battles"][label] += 1

    our_rows   = bpr_adr_sort([r for a, p in our_players.items()   if (r := calc_stats(a, p))])
    enemy_rows = bpr_adr_sort([r for a, p in enemy_players.items() if (r := calc_stats(a, p))])

    return our_rows, enemy_rows, processed, errors


# ── routes ──────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory("static", "index.html")


@app.route("/api/process", methods=["POST"])
def process():
    data   = request.json
    folder = data.get("folder", "").strip()
    if not folder or not os.path.isdir(folder):
        return jsonify({"error": "Invalid folder path"}), 400

    our, enemy, processed, errors = process_replays(folder)
    return jsonify({
        "our_team":   our,
        "enemy_team": enemy,
        "processed":  processed,
        "errors":     errors,
    })


@app.route("/api/upload", methods=["POST"])
def upload():
    """Accept uploaded replay files from the browser folder picker."""
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "No files uploaded"}), 400

    tmp_dir = tempfile.mkdtemp(prefix="blitz_replays_")
    saved   = 0
    for f in files:
        if f.filename.endswith(".wotbreplay"):
            # strip any path prefix the browser may include
            fname = os.path.basename(f.filename)
            f.save(os.path.join(tmp_dir, fname))
            saved += 1

    if saved == 0:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        return jsonify({"error": "No .wotbreplay files found in upload"}), 400

    our, enemy, processed, errors = process_replays(tmp_dir)
    shutil.rmtree(tmp_dir, ignore_errors=True)

    return jsonify({
        "our_team":   our,
        "enemy_team": enemy,
        "processed":  processed,
        "errors":     errors,
    })


@app.route("/api/export", methods=["POST"])
def export():
    data    = request.json
    our     = data.get("our_team", [])
    enemy   = data.get("enemy_team", [])

    wb = Workbook()

    KEYS = [
        "account_id", "nickname", "battles",
        "HT", "MT", "LT", "TD", "main_tank",
        "ADR", "frags", "KPR", "DE", "assist", "blocked",
        "shots", "hits", "pens",
        "AccH", "AccP",
        "iPoints", "sPoints",
        "Firepower", "AIM", "Support", "Supremacy", "BPR"
    ]
    HEADERS = [
        "account_id", "nickname", "battles",
        "HT", "MT", "LT", "TD", "main_tank",
        "ADR", "Frags", "KPR", "DE", "Assist", "Blocked",
        "shots", "hits", "pens",
        "AccH (%)", "AccP (%)",
        "iPoints", "sPoints",
        "Firepower", "AIM", "Support", "Supremacy", "BPR 2.0"
    ]
    COL_W = [14,22,9,6,6,6,6,18,8,7,7,7,9,9,8,8,8,10,10,9,9,11,9,10,11,9]

    hdr_fill  = PatternFill("solid", fgColor="FF5500")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    even_fill = PatternFill("solid", fgColor="1A1A1A")
    odd_fill  = PatternFill("solid", fgColor="141414")
    dat_font  = Font(color="E0E0E0", name="Calibri", size=10)
    ctr       = Alignment(horizontal="center", vertical="center")
    btm       = Border(bottom=Side(style="thin", color="2A2A2A"))

    def write_sheet(ws, rows, title_color):
        ws.append(HEADERS)
        hf = PatternFill("solid", fgColor=title_color)
        for cell in ws[1]:
            cell.fill = hf; cell.font = hdr_font; cell.alignment = ctr
        for i, player in enumerate(rows):
            ws.append([player.get(k, "") for k in KEYS])
            fill = even_fill if i % 2 == 0 else odd_fill
            for cell in ws[i + 2]:
                cell.fill = fill; cell.font = dat_font
                cell.alignment = ctr; cell.border = btm
        for i, w in enumerate(COL_W, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w
        ws.row_dimensions[1].height = 22

    ws_our = wb.active
    ws_our.title = "Our Team"
    write_sheet(ws_our, our, "FF5500")

    ws_enemy = wb.create_sheet("Enemy Team")
    write_sheet(ws_enemy, enemy, "444444")

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    return send_file(
        tmp.name,
        as_attachment=True,
        download_name="results.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    os.makedirs("output", exist_ok=True)
    app.run(debug=True, port=5000)